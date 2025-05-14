
import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

st.set_page_config(page_title="교정일 통합관리 시스템", layout="wide")
st.title("📋 연도·월별 차기점검일 통합관리 앱")

uploaded_file = st.file_uploader("📤 Excel 파일을 업로드하세요", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    all_data = []

    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet, skiprows=2)
            target_cols = [col for col in df.columns if '차기점검일' in str(col)]
            for col in target_cols:
                df[col] = pd.to_datetime(df[col], errors="coerce")
                df["연도"] = df[col].dt.year
                df["월"] = df[col].dt.month
                df["차기점검일"] = df[col]
                df["차기점검일_열"] = col
                df["시트명"] = sheet

                target_keywords = ["차종", "품번", "구분", "단수", "Gauge", "관리번호", "판정", "비고"]
                keep_cols = [c for c in df.columns if any(k in str(c) for k in target_keywords)]

                filtered = df[["시트명", "차기점검일", "연도", "월"] + keep_cols]
                all_data.append(filtered)
        except Exception:
            continue

    if all_data:
        merged_df = pd.concat(all_data, ignore_index=True)
        merged_df.dropna(subset=["차기점검일"], inplace=True)

        years = sorted(merged_df["연도"].dropna().astype(int).unique())
        selected_year = st.selectbox("연도 선택", years)
        months = sorted(merged_df[merged_df["연도"] == selected_year]["월"].dropna().astype(int).unique())
        selected_month = st.selectbox("월 선택", months)

        result_df = merged_df[
            (merged_df["연도"] == selected_year) & 
            (merged_df["월"] == selected_month)
        ].reset_index(drop=True)

        st.markdown(f"🔎 **{selected_year}년 {selected_month}월 차기점검 대상: {len(result_df)}건**")
        st.dataframe(result_df, use_container_width=True)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.book.active
            for column_cells in worksheet.columns:
                max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
                adjusted_width = max_length + 12
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        buffer.seek(0)
        st.download_button(
            label=f"📥 {selected_year}-{selected_month:02d} 교정대상 다운로드",
            data=buffer,
            file_name=f"교정대상_{selected_year}-{selected_month:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.subheader("📊 선택 연도 내 월별 교정 대상 건수")
        chart_data = merged_df[merged_df["연도"] == selected_year]["월"].value_counts().sort_index()
        st.bar_chart(chart_data)
    else:
        st.warning("⚠ '차기점검일' 열을 가진 유효한 시트가 없습니다.")
